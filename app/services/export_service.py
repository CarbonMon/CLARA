import pandas as pd
import io
from typing import List, Dict, Any
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def create_excel_file(data: List[Dict[str, Any]], filename: str) -> io.BytesIO:
    """
    Create an Excel file from a list of dictionaries with advanced formatting
    
    Args:
        data: List of dictionaries to convert to Excel
        filename: Name of the Excel file
        
    Returns:
        BytesIO object containing the Excel file
    """
    df = pd.DataFrame(data)
    excel_file = io.BytesIO()
    
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    
    # Apply advanced formatting
    excel_file.seek(0)
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # Define column widths (in Excel units)
    column_widths = {
        "Title": 40,
        "PMID": 10,
        "Full Text Link": 30,
        "Analysis Source": 20,
        "Subject of Study": 15,
        "Disease State": 25,
        "Number of Subjects Studied": 10,
        "Type of Study": 20,
        "Type of Study 2": 15,
        "Study Design": 30,
        "Intervention": 30,
        "Intervention Dose": 30,
        "Intervention Dosage Form": 20,
        "Control": 25,
        "Primary Endpoint": 35,
        "Primary Endpoint Result": 35,
        "Secondary Endpoints": 35,
        "Safety Endpoints": 25,
        "Results Available": 10,
        "Primary Endpoint Met": 10,
        "Statistical Significance": 25,
        "Clinical Significance": 25,
        "Main Author": 20,
        "Other Authors": 40,
        "Journal Name": 30,
        "Date of Publication": 15,
        "Error": 10,
        "Filename": 25,
        "Search Query": 50
    }
    
    # Format headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Set column widths
    for col_idx, column in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        if column in column_widths:
            ws.column_dimensions[col_letter].width = column_widths[column]
        else:
            ws.column_dimensions[col_letter].width = 15  # Default width
    
    # Create a table with the data
    table_ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"
    tab = Table(displayName="ResultsTable", ref=table_ref)
    
    # Add a professional style
    style = TableStyleInfo(
        name="TableStyleMedium9", 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True, 
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    
    # Add the table to the worksheet
    ws.add_table(tab)

    # Add AI disclaimer at the bottom of the results
    disclaimer_row = len(df) + 3  # Skip a row after the table
    disclaimer_text = "Notice: Results are AI-generated. Please verify findings with primary sources."
    ws.cell(row=disclaimer_row, column=1, value=disclaimer_text)
    ws.merge_cells(start_row=disclaimer_row, start_column=1, end_row=disclaimer_row, end_column=len(df.columns))
    disclaimer_cell = ws.cell(row=disclaimer_row, column=1)
    disclaimer_cell.alignment = Alignment(horizontal='left', vertical='center')
    disclaimer_cell.font = Font(italic=True, size=9, color="666666")

    # Add metadata row
    metadata_row = len(df) + 4
    metadata_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Tool: CLARA Clinical Research Analysis Tool"
    ws.cell(row=metadata_row, column=1, value=metadata_text)
    ws.merge_cells(start_row=metadata_row, start_column=1, end_row=metadata_row, end_column=len(df.columns))
    metadata_cell = ws.cell(row=metadata_row, column=1)
    metadata_cell.alignment = Alignment(horizontal='left', vertical='center')
    metadata_cell.font = Font(size=8, color="999999")

    # Save the formatted workbook
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file
